import datetime
from django.db.models import Q
from django.http import HttpResponse
from django.views.generic import View, ListView, DetailView, TemplateView
from django.shortcuts import render
from django.contrib.auth.mixins import LoginRequiredMixin
from import_export import resources
from openpyxl import Workbook

from employee.model.employee import Employee
from pure_pagination.mixins import PaginationMixin

from directory.models import City, District, Country, EducationType
from directory.models import Language as DLanguage
from root.mixins import IsSuperUserMixin


class Employees(IsSuperUserMixin, PaginationMixin, ListView):
    paginate_by = 10
    context_object_name = 'employees'
    template_name = 'root/employee/employees.html'

    def get_queryset(self):
        data = self.request.GET
        fullname, age, gender = data.get("fullname"), data.get("age"), data.get("gender")
        language, education, category = data.get("language"), data.get("education"), data.get("category")
        qs = Employee.objects.all()
        if fullname:
            qs = qs.filter(
                Q(full_name_en__icontains=fullname) |
                Q(full_name_ru__icontains=fullname)
            )
        if age:
            ages = age.split('-')
            ages = [int(i) for i in ages if i.isdigit()]
            if len(ages) == 2:
                td = datetime.datetime.today()
                y1 = datetime.date(td.year - ages[0], 1, 1)
                y2 = datetime.date(td.year - ages[1], 1, 1)
                qs = qs.filter(birth_date__lte=y1, birth_date__gte=y2)
            elif len(ages) == 1:
                year = datetime.datetime.today().year - int(age)
                qs = qs.filter(birth_date__year=year)
            # year = datetime.datetime.today().year - int(age)
            # qs = qs.filter(birth_date__year=year)
        if gender:
            qs = qs.filter(gender=gender)
        if language:
            qs = qs.filter(language__language_id__in=list(language))
        if education:
            qs = qs.filter(education__type_id__in=list(education))
        if category:
            qs = qs.filter(score__category=category)

        register_number = data.get('register_number')
        if register_number:
            qs = qs.filter(register_number__icontains=register_number)
        passport_serial = data.get('passport_serial')
        if passport_serial:
            qs = qs.filter(passport_serial__icontains=passport_serial)
        age = data.get('age')
        if age:
            if age.isdigit():
                qs = qs.filter(birth_date__year=datetime.date.today().year - int(age))
        fatness = data.get('fatness')
        if fatness:
            qs = qs.filter(fatness=fatness)
        education = data.get('education')
        if education:
            qs = qs.filter(education__type__id=education)
        criminal = data.get('criminal')
        if criminal:
            qs = qs.filter(criminal_issue__isnull=False)
        military_duty = data.get('military_duty')
        if military_duty:
            qs = qs.filter(military_duty=military_duty)
        family_status = data.get('family_status')
        if family_status:
            qs = qs.filter(family__status=family_status)

        fullname = data.get('fullname')
        if fullname:
            qs = qs.filter(
                Q(full_name_ru__icontains=fullname) |
                Q(full_name_en__icontains=fullname)
            )
        inn = data.get('inn')
        if inn:
            qs = qs.filter(inn__icontains=inn)
        height = data.get('height')
        if height:
            qs = qs.filter(height=height)
        weight = data.get('weight')
        if weight:
            qs = qs.filter(weight=weight)
        vision_l = data.get('vision_l')
        if vision_l:
            qs = qs.filter(vision_l=vision_l)
        vision_r = data.get('vision_r')
        if vision_r:
            qs = qs.filter(vision_r=vision_r)
        language = data.get('language')
        if language:
            qs = qs.filter(language__language__id=language)
        language_level = data.get('language_level')
        if language_level:
            qs = qs.filter(language__level=language_level)

        reward = data.get('reward')
        if reward:
            qs = qs.filter(reward__isnull=reward)
        country = data.get('country')
        if country:
            qs = qs.filter(countries__country_id=country)
        level = data.get('level')
        if level:
            if level == 'is_employee':
                qs = qs.filter(is_employee=True)
            elif level == 'is_student':
                qs = qs.filter(is_student=True)
            elif level == 'is_young_talent':
                qs = qs.filter(is_young_talent=True)
        gender = data.get('gender')
        if gender:
            qs = qs.filter(gender=gender)
        phone = data.get('phone')
        if phone:
            qs = qs.filter(phone__icontains=phone)
        experience = data.get('experience')
        # if experience:
        #     if experience.isdigit():
        #         total = 0
        #         for i in
        wages = data.get('wages')
        if wages:
            qs = qs.filter(wages=wages)
        category = data.get('category')
        if category:
            qs = qs.filter(score__category=category)
        return qs.order_by('-register_number')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        data = self.request.GET
        fullname, age, gender = data.get("fullname"), data.get("age"), data.get("gender")
        language, education, category = data.get("language"), data.get("education"), data.get("category")
        if fullname:
            context['fullname'] = fullname
        if age:
            context['age'] = age
        if gender:
            context['gender'] = gender
        if language:
            if language.isdigit():
                context['language'] = int(language)
        if education:
            if education.isdigit():
                context['education'] = int(education)
        if category:
            context['category'] = category

        context['educations'] = EducationType.objects.all()
        context['languages'] = DLanguage.objects.all()

        return context


class EmployeeDetail(IsSuperUserMixin, DetailView):
    pk_url_kwarg = 'id'
    model = Employee
    template_name = 'root/employee/employee_detail.html'
    context_object_name = 'employee'


class EmployeeDetailExportPDF(IsSuperUserMixin, DetailView):
    pk_url_kwarg = 'id'
    model = Employee
    template_name = 'root/employee/personal_card.html'
    context_object_name = 'employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['countries'] = Country.objects.all()
        return context


class EmployeeUpdateOP1(IsSuperUserMixin, DetailView):
    pk_url_kwarg = 'id'
    model = Employee
    template_name = 'root/employee/employee_update_1.html'
    context_object_name = 'employee'


class EmployeeUpdateOP2(IsSuperUserMixin, DetailView):
    pk_url_kwarg = 'id'
    model = Employee
    template_name = 'root/employee/employee_update_2.html'
    context_object_name = 'employee'


class EmployeeUpdateOP3(IsSuperUserMixin, DetailView):
    pk_url_kwarg = 'id'
    model = Employee
    template_name = 'root/employee/employee_update_3.html'
    context_object_name = 'employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["cities"] = City.objects.all()
        context["educations"] = EducationType.objects.all()
        context["districts"] = District.objects.all()
        context["languages"] = DLanguage.objects.all()
        return context


class EmployeeUpdateOP4(IsSuperUserMixin, DetailView):
    pk_url_kwarg = 'id'
    model = Employee
    template_name = 'root/employee/employee_update_4.html'
    context_object_name = 'employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['countries'] = Country.objects.all()
        # context['emp_countries'] = ', '.join(str(i) for i in list(Country.objects.all().values_list('id', flat=True)))
        context['emp_countries'] = list(self.object.countries.all().values_list('country_id', flat=True))
        return context


class EmployeeUpdateRequests(IsSuperUserMixin, View):
    template_name = "root/employee_update_request.html"

    def get(self, request):
        return render(request, self.template_name, {})


class EmployeeTranslation1(IsSuperUserMixin, TemplateView):
    template_name = "root/employee/employee_translation_1.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["employee"] = Employee.objects.get(id=self.kwargs.get("id"))
        return context


class EmployeeTranslation3(IsSuperUserMixin, DetailView):
    template_name = "root/employee/employee_translation_3.html"
    pk_url_kwarg = 'id'
    queryset = Employee.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["employee"] = Employee.objects.get(id=self.kwargs.get("id"))
        context["cities"] = City.objects.all()
        context["educations"] = EducationType.objects.all()
        context["districts"] = District.objects.all()
        context["languages"] = DLanguage.objects.all()
        context['emp_countries'] = list(self.object.countries.all().values_list('country_id', flat=True))
        return context


class EmployeeTranslation4(IsSuperUserMixin, DetailView):
    template_name = "root/employee/employee_translation_4.html"
    pk_url_kwarg = 'id'
    queryset = Employee.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["employee"] = Employee.objects.get(id=self.kwargs.get("id"))
        context["countries"] = Country.objects.all()
        context['emp_countries'] = list(self.object.countries.all().values_list('country_id', flat=True))
        return context


class ChangedEmployeesListView(IsSuperUserMixin, ListView):
    template_name = 'root/employee/changed_employees.html'
    paginate_by = 10
    context_object_name = 'employees'

    def get_queryset(self):
        return Employee.objects.filter(
            Q(employeechanges__isnull=False) |
            Q(education__educationchanges__isnull=False) |
            Q(language__languagechanges__isnull=False) |
            Q(army__armychanges__isnull=False) |
            Q(relative__relativechanges__isnull=False) |
            Q(reward__rewardchanges__isnull=False) |
            Q(experience__experiencechanges__isnull=False) |
            Q(family__familychanges__isnull=False)
        ).distinct()


class ChangedEmployeeDetailView(IsSuperUserMixin, DetailView):
    template_name = 'root/employee/changed_employee.html'
    context_object_name = 'employee'
    pk_url_kwarg = 'id'
    model = Employee


# statistics
class EmployeeStatistics(IsSuperUserMixin, ListView):
    template_name = 'root/employee/employee_statistics.html'
    context_object_name = 'employees'

    def get_queryset(self):
        data = self.request.GET
        region = data.get('region')
        family = data.get('family')
        fatness = data.get('fatness')
        to_university = data.get('to_university')

        education = data.get('education')
        language = data.get('language')
        height = data.get('height')
        weight = data.get('weight')
        country = data.get('country')
        category = data.get('category')
        psycholog = data.get('psycholog')
        wages = data.get('wages')
        qs = Employee.objects.all()

        if region:
            qs = qs.filter(
                Q(birth_place_ru__icontains=region) |
                Q(birth_place_en__icontains=region)
            )
        if family:
            qs = qs.filter(family__status=family)
        if fatness:
            fatness = fatness.split('-')
            if len(fatness) == 2:
                if fatness[0].startswith('0'):
                    fatness[0] = fatness[0][1:]
                    qs = qs.filter(fatness__range=fatness)
                else:
                    qs = qs.filter(fatness__range=fatness)
            elif len(fatness) == 1:
                qs = qs.filter(fatness__gte=fatness[0])
        if to_university:
            qs = qs.filter(is_ready_for_university=True if to_university == 'true' else False)
        if education:
            qs = qs.filter(education__type_id=education)
        if height:
            height = height.split('-')
            if len(height) == 2:
                if height[0].startswith('0'):
                    height[0] = height[0][1:]
                    qs = qs.filter(height__range=height)
                else:
                    qs = qs.filter(height__range=height)
            elif len(height) == 1:
                qs = qs.filter(height__gte=height[0])
        if weight:
            weight = weight.split('-')
            if len(weight) == 2:
                if weight[0].startswith('0'):
                    weight[0] = weight[0][1:]
                    qs = qs.filter(weight__range=weight)
                else:
                    qs = qs.filter(weight__range=weight)
            elif len(weight) == 1:
                qs = qs.filter(weight__gte=weight[0])
        if country:
            qs = qs.filter(countries__country_id=country)
        if psycholog:
            qs = qs.filter(psycholog=psycholog)
        if language:
            qs = qs.filter(language__language_id=language)
        if category:
            qs = qs.filter(score__category=category)
        if wages:
            qs = qs.filter(wages=wages)
        return qs

    def get(self, request, *args, **kwargs):
        if request.GET.get('download'):
            data = self.request.GET
            region = data.get('region')
            family = data.get('family')
            fatness = data.get('fatness')
            to_university = data.get('to_university')

            education = data.get('education')
            language = data.get('language')
            height = data.get('height')
            weight = data.get('weight')
            country = data.get('country')
            category = data.get('category')
            psycholog = data.get('psycholog')
            wages = data.get('wages')
            qs = Employee.objects.all()

            if region:
                qs = qs.filter(
                    Q(birth_place_ru__icontains=region) |
                    Q(birth_place_en__icontains=region)
                )
            if family:
                qs = qs.filter(family__status=family)
            if fatness:
                fatness = fatness.split('-')
                if len(fatness) == 2:
                    if fatness[0].startswith('0'):
                        fatness[0] = fatness[0][1:]
                        qs = qs.filter(fatness__range=fatness)
                    else:
                        qs = qs.filter(fatness__range=fatness)
                elif len(fatness) == 1:
                    qs = qs.filter(fatness__gte=fatness[0])
            if to_university:
                qs = qs.filter(is_ready_for_university=True if to_university == 'true' else False)
            if education:
                qs = qs.filter(education__type_id=education)
            if height:
                height = height.split('-')
                if len(height) == 2:
                    if height[0].startswith('0'):
                        height[0] = height[0][1:]
                        qs = qs.filter(height__range=height)
                    else:
                        qs = qs.filter(height__range=height)
                elif len(height) == 1:
                    qs = qs.filter(height__gte=height[0])
            if weight:
                weight = weight.split('-')
                if len(weight) == 2:
                    if weight[0].startswith('0'):
                        weight[0] = weight[0][1:]
                        qs = qs.filter(weight__range=weight)
                    else:
                        qs = qs.filter(weight__range=weight)
                elif len(weight) == 1:
                    qs = qs.filter(weight__gte=weight[0])
            if country:
                qs = qs.filter(countries__country_id=country)
            if psycholog:
                qs = qs.filter(psycholog=psycholog)
            if language:
                qs = qs.filter(language__language_id=language)
            if category:
                qs = qs.filter(score__category=category)
            if wages:
                qs = qs.filter(wages=wages)
            return export_to_xls(qs)
        return super(EmployeeStatistics, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        if self.request.GET.get('download'):
            print(11)
            export_to_xls(self.request)

        context = super().get_context_data(**kwargs)
        context['languages'] = DLanguage.objects.all()
        context['regions'] = City.objects.all()
        context['educations'] = EducationType.objects.all()
        context['countries'] = Country.objects.all()

        data = self.request.GET
        region = data.get('region')
        if region: context['region'] = region
        family = data.get('family')
        if family: context['family'] = family
        fatness = data.get('fatness')
        if fatness: context['fatness'] = fatness
        to_university = data.get('to_university')
        if to_university: context['to_university'] = to_university
        education = data.get('education')
        if education: context['education'] = int(education) if education.isdigit() else 0
        language = data.get('language')
        if language: context['language'] = int(language) if language.isdigit() else 0
        height = data.get('height')
        if height: context['height'] = height
        weight = data.get('weight')
        if weight: context['weight'] = weight
        country = data.get('country')
        if country: context['country'] = int(country) if country.isdigit() else 0
        category = data.get('category')
        if category: context['category'] = category
        psycholog = data.get('psycholog')
        if psycholog: context['psycholog'] = psycholog
        wages = data.get('wages')
        if wages: context['wages'] = wages
        return context


class EmployeeExportXLS(resources.ModelResource):

    class Meta:
        model = Employee
        fields = (
            'full_name_ru',
            'full_name_en',
            'birth_date',
            'birth_place_ru',
            'living_address_ru',
            'inn',
        )


def export_to_xls(qs):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(
        date=datetime.datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Аппликанты'

    # Define the titles for columns
    columns = [
        'ID',
        'Ф.И.О (русс)',
        'Ф.И.О (англ)',
        'Дата рождения',
        'Место рождения',
        'Место проживания',
    ]
    row_num = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    # Iterate through all movies
    for e in qs:
        row_num += 1
        # Define the data for each cell in the row
        row = [
            e.id,
            e.full_name_ru,
            e.full_name_en,
            e.birth_date,
            e.birth_place_ru,
            e.living_address_ru,
        ]
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response
